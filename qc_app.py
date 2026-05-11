import streamlit as st
import pandas as pd
import io, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

st.set_page_config(page_title="QC Report Generator", page_icon="🔍", layout="centered")
st.title("Quality Control Report Generator")
st.markdown("Upload an Awitec inspection file (.xlsx) to generate a structured QC report.")

# ── Colors ────────────────────────────────────────────────────────────────────
C_DARK      = "1F3864"
C_MID       = "2E75B6"
C_OK_HDR    = "A9D18E"
C_NOK_HDR   = "FF9999"
C_OK_DATA   = "E2EFDA"
C_NOK_DATA  = "FFE0E0"
C_RW_DATA   = "CCFFCC"
C_TOTAL_ROW = "BDD7EE"
C_ODD       = "EBF3FB"
C_WHITE     = "FFFFFF"
C_HIGHLIGHT = "FFF2CC"
INSP_COLORS = ["4472C4", "ED7D31", "2E86C1", "8E44AD"]
RW_COLOR    = "70AD47"

def fl(h):   return PatternFill("solid", start_color=h, fgColor=h)
def th():
    s = Side(style="thin", color="000000")
    return Border(left=s, right=s, top=s, bottom=s)
def ctr(wrap=True): return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
def lft():          return Alignment(horizontal="left",   vertical="center", wrap_text=False)
def fnt(bold=False, sz=10, color="000000"):
    return Font(name="Arial", bold=bold, size=sz, color=color)

# ── Dynamic column detection ──────────────────────────────────────────────────
def detect_columns(df, group_row_idx, subheader_row_idx):
    """
    Returns a list of inspection/rework dicts:
      { name, label, col_ok, col_nok, is_rework }
    Inspection 2 (handling) is skipped.
    Also returns total_col index.
    """
    group_row = df.iloc[group_row_idx]
    sub_row   = df.iloc[subheader_row_idx]

    total_col = next(
        (j for j, v in enumerate(sub_row) if "Pieces to check" in str(v)),
        None
    )

    detections = []
    for j, v in enumerate(group_row):
        if not pd.notna(v) or not str(v).strip():
            continue
        label = str(v).strip()

        # Skip Inspection 2 (handling / extra packaging)
        if re.search(r"Inspection\s*2\b", label, re.I):
            continue
        # Skip Inspection 4 (not used)
        if re.search(r"Inspection\s*4\b", label, re.I):
            continue

        is_rework = bool(re.search(r"rework|reworking", label, re.I))
        is_insp   = bool(re.search(r"Inspection\s*[13]", label, re.I))

        if not (is_rework or is_insp):
            continue

        # Find OK and NOK sub-columns starting at j
        col_ok = col_nok = None
        for offset in range(0, 4):
            sub_val = str(sub_row.iloc[j + offset]) if j + offset < len(sub_row) else ""
            if "Pieces good" in sub_val:
                col_ok = j + offset
            elif "Pieces not good" in sub_val:
                col_nok = j + offset
            if col_ok is not None and col_nok is not None:
                break

        if col_ok is None or col_nok is None:
            continue

        # Clean display label
        clean = re.sub(r"^(Inspection\s*\d+|Reworking?)\s*", "", label, flags=re.I).strip().strip("()")
        if not clean:
            clean = label

        detections.append({
            "name":      label,
            "label":     clean,
            "col_ok":    col_ok,
            "col_nok":   col_nok,
            "is_rework": is_rework,
        })

    return detections, total_col


def is_valid_partno(val):
    """Filter out non-part-number rows (e.g. free-text notes)."""
    s = str(val).strip()
    if not s or s.lower() in ["nan", "none", "divers"]:
        return False
    # Valid part numbers: contain digits and letters, no long spaces
    if len(s) > 30:
        return False
    if " " in s and not re.match(r"^[A-Z0-9]{3,}-", s):
        return False
    return bool(re.search(r"[A-Z0-9]", s, re.I))


# ── File parsing ──────────────────────────────────────────────────────────────
def parse_file(uploaded_file):
    df = pd.read_excel(uploaded_file, sheet_name="Order Awitec", header=None)

    subheader_row = next(
        (i for i, row in df.iterrows() if any("Pieces to check" in str(v) for v in row.values)),
        None
    )
    if subheader_row is None:
        return None, "Could not find header row ('Pieces to check')."

    group_row = subheader_row - 1
    total_row_idx = next(
        (i for i, row in df.iterrows() if any("Total Quantity" in str(v) for v in row.values)),
        None
    )
    if total_row_idx is None:
        return None, "Could not find total row ('Total Quantity')."

    # Metadata
    meta = {"order": "", "material": "", "description": ""}
    for i, row in df.iloc[:subheader_row].iterrows():
        for j, v in enumerate(row):
            if not pd.notna(v): continue
            key = str(v).strip()
            val_col = j + 4
            val = df.iloc[i, val_col] if val_col < df.shape[1] else None
            if "Order/Job Number" in key and pd.notna(val):
                meta["order"] = str(val).strip()
            elif "Material No." in key and pd.notna(val):
                meta["material"] = str(val).strip()
            elif "Order description" in key and pd.notna(val):
                meta["description"] = str(val).strip()

    # Detect columns dynamically
    detections, total_col = detect_columns(df, group_row, subheader_row)
    if total_col is None:
        return None, "Could not detect total quantity column."
    if not detections:
        return None, "No inspection columns detected."

    # Find Charge column (in group-header row, label "Charge")
    charge_col = next(
        (j for j, v in enumerate(df.iloc[group_row]) if "charge" in str(v).lower()),
        None
    )

    # Load data
    base_cols = [1, 2, total_col]
    if charge_col is not None:
        base_cols.append(charge_col)
    all_cols = base_cols + [d["col_ok"] for d in detections] + [d["col_nok"] for d in detections]
    all_cols = sorted(set(all_cols))

    data = df.iloc[subheader_row+1:total_row_idx, all_cols].copy()
    data.columns = all_cols

    rename = {1: "date", 2: "partno", total_col: "total"}
    if charge_col is not None:
        rename[charge_col] = "charge"
    for d in detections:
        key = re.sub(r"[^a-z0-9]", "_", d["label"].lower())[:20]
        rename[d["col_ok"]]  = key + "_ok"
        rename[d["col_nok"]] = key + "_nok"
        d["key_ok"]  = key + "_ok"
        d["key_nok"] = key + "_nok"

    data = data.rename(columns=rename)
    data["date"]   = pd.to_datetime(data["date"],   errors="coerce")
    data["partno"] = data["partno"].astype(str).str.strip()
    if "charge" in data.columns:
        data["charge"] = data["charge"].astype(str).str.strip().replace("nan", "")

    num_cols = [c for c in data.columns if c not in ["date", "partno", "charge"]]
    for c in num_cols:
        data[c] = pd.to_numeric(data[c], errors="coerce").fillna(0).astype(int)

    # Filter valid rows
    data = data[data["date"].notna()]
    data = data[data["partno"].apply(is_valid_partno)]
    data = data[data["total"] > 0]

    # Detect which inspections/rework have actual data
    active = [d for d in detections if data[d["key_nok"]].sum() + data[d["key_ok"]].sum() > 0]
    active_inspections = [d for d in active if not d["is_rework"]]
    active_rework      = [d for d in active if d["is_rework"]]

    part_numbers = sorted(data["partno"].unique())
    multi_part   = len(part_numbers) > 1

    # Per-part rework detection
    rework_parts = []
    for d in active_rework:
        rw_parts = data[data[d["key_ok"]] + data[d["key_nok"]] > 0]["partno"].unique()
        rework_parts.extend(rw_parts)
    rework_parts = sorted(set(rework_parts))

    return {
        "data": data,
        "meta": meta,
        "inspections": active_inspections,
        "rework": active_rework,
        "rework_parts": rework_parts,
        "part_numbers": part_numbers,
        "multi_part": multi_part,
        "has_charge": "charge" in data.columns,
    }, None


# ── Report generation ─────────────────────────────────────────────────────────
def generate_report(info):
    data        = info["data"]
    meta        = info["meta"]
    inspections = info["inspections"]
    rework_list = info["rework"]
    rework_parts = info["rework_parts"]
    multi_part  = info["multi_part"]
    has_charge  = info.get("has_charge", False)

    order_id = meta.get("order", "Unknown")
    material = meta.get("material", "")

    group_cols = ["date", "partno"] + (["charge"] if has_charge else [])
    agg = {"total": ("total", "sum")}
    for d in inspections + rework_list:
        agg[d["key_ok"]]  = (d["key_ok"],  "sum")
        agg[d["key_nok"]] = (d["key_nok"], "sum")

    by_day  = data.groupby(group_cols).agg(**agg).reset_index()
    by_day  = by_day.sort_values(group_cols)
    by_part = data.groupby("partno").agg(**{k: v for k, v in agg.items()}).reset_index().sort_values("partno") if multi_part else None

    wb = Workbook()

    # ── SHEET 1: Daily Report ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Daily Report"
    ws.sheet_view.showGridLines = False

    # Build column list
    cols = [{"key":"date",   "label":"Date",     "width":13},
            {"key":"partno", "label":"Part No.", "width":16}]
    if has_charge:
        cols.append({"key":"charge", "label":"Charge", "width":14})
    cols.append({"key":"total", "label":"Total\nQty.", "width":12})

    for ci, d in enumerate(inspections):
        color = INSP_COLORS[ci % len(INSP_COLORS)]
        cols.append({"key":d["key_ok"],  "label":"OK",  "width":11, "group":f"INSP – {d['label']}", "color":color})
        cols.append({"key":d["key_nok"], "label":"NOK", "width":11, "group":f"INSP – {d['label']}", "color":color})

    for d in rework_list:
        rw_label = d["label"] or "Rework"
        cols.append({"key":d["key_ok"],  "label":"OK",  "width":12, "group":f"REWORK – {rw_label}", "color":RW_COLOR})
        cols.append({"key":d["key_nok"], "label":"NOK", "width":12, "group":f"REWORK – {rw_label}", "color":RW_COLOR})

    cols.append({"key":"total_ok",  "label":"Total\nOK",  "width":11, "group":"TOTALS", "color":C_MID})
    cols.append({"key":"total_nok", "label":"Total\nNOK", "width":11, "group":"TOTALS", "color":C_MID})
    cols.append({"key":"nok_rate",  "label":"NOK\nRate %","width":11})
    cols.append({"key":"remarks",   "label":"Remarks",    "width":28})

    n_cols = len(cols)
    for i, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = col["width"]

    # Row 1: Title
    ws.row_dimensions[1].height = 32
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    tc = ws.cell(row=1, column=1)
    tc.value = f"QUALITY CONTROL REPORT  |  Order: {order_id}  |  {material}"
    tc.font  = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    tc.fill  = fl(C_DARK); tc.alignment = ctr()
    for c in range(1, n_cols+1): ws.cell(row=1, column=c).border = th()

    # Row 2: Group headers / Row 3: Sub-headers
    ws.row_dimensions[2].height = 26
    ws.row_dimensions[3].height = 32

    # Static columns span rows 2-3
    static_count = 3 + (1 if has_charge else 0)  # date, partno, [charge], total
    for ci in range(1, static_count+1):
        ws.merge_cells(start_row=2, start_column=ci, end_row=3, end_column=ci)
        for r in [2,3]: ws.cell(row=r, column=ci).border = th()
        c = ws.cell(row=2, column=ci)
        c.value = cols[ci-1]["label"]
        c.font  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill  = fl(C_DARK); c.alignment = ctr()

    # Group headers for paired columns
    ci = static_count + 1
    while ci <= n_cols:
        col = cols[ci-1]
        key = col["key"]
        if key in ["nok_rate", "remarks"]:
            ws.merge_cells(start_row=2, start_column=ci, end_row=3, end_column=ci)
            for r in [2,3]: ws.cell(row=r, column=ci).border = th()
            c = ws.cell(row=2, column=ci)
            c.value = col["label"]
            c.fill  = fl("FFE699" if key=="nok_rate" else C_DARK)
            c.font  = Font(name="Arial", bold=True, size=10,
                           color="000000" if key=="nok_rate" else "FFFFFF")
            c.alignment = ctr()
            ci += 1
        elif "group" in col:
            grp  = col["group"]
            span = sum(1 for c2 in cols[ci-1:] if c2.get("group") == grp)
            ws.merge_cells(start_row=2, start_column=ci, end_row=2, end_column=ci+span-1)
            for cx in range(ci, ci+span): ws.cell(row=2, column=cx).border = th()
            hc = ws.cell(row=2, column=ci)
            hc.value = grp
            hc.font  = Font(name="Arial", bold=True, size=9, color="FFFFFF")
            hc.fill  = fl(col["color"]); hc.alignment = ctr()
            for c2 in cols[ci-1:ci-1+span]:
                sub_ci = cols.index(c2)+1
                sub = ws.cell(row=3, column=sub_ci)
                is_ok = "ok" in c2["key"] and "nok" not in c2["key"]
                sub.value = c2["label"]
                sub.font  = Font(name="Arial", bold=True, size=9)
                sub.fill  = fl(C_OK_HDR if is_ok else C_NOK_HDR)
                sub.alignment = ctr(); sub.border = th()
            ci += span
        else:
            ci += 1

    # Data rows
    DATA_START = 4
    nok_keys    = [d["key_nok"] for d in inspections]
    rw_ok_keys  = [d["key_ok"]  for d in rework_list]

    def total_col_ci():
        return next(i+1 for i,c in enumerate(cols) if c["key"]=="total")
    def tnok_col_ci():
        return next(i+1 for i,c in enumerate(cols) if c["key"]=="total_nok")

    for idx, row in enumerate(by_day.itertuples(index=False)):
        r   = DATA_START + idx
        ws.row_dimensions[r].height = 18
        rd  = row._asdict()
        pno = rd.get("partno", "")
        is_rw_part = pno in rework_parts
        total_nok  = sum(rd.get(k, 0) for k in nok_keys) - sum(rd.get(k, 0) for k in rw_ok_keys)
        has_nok    = total_nok > 0
        bg = C_HIGHLIGHT if (is_rw_part and multi_part) else (C_ODD if idx%2==0 else C_WHITE)

        for ci, col in enumerate(cols, 1):
            c = ws.cell(row=r, column=ci)
            key = col["key"]

            if key == "date":
                c.value = rd["date"].date()
                c.number_format = "DD.MM.YYYY"
                c.font = fnt(); c.fill = fl(bg); c.border = th(); c.alignment = ctr()

            elif key == "partno":
                c.value = pno
                c.font  = fnt(bold=is_rw_part); c.fill = fl(bg); c.border = th(); c.alignment = ctr()

            elif key == "charge":
                c.value = rd.get("charge", "")
                c.font  = fnt(); c.fill = fl(bg); c.border = th(); c.alignment = ctr()

            elif key == "total":
                c.value = int(rd["total"])
                c.number_format = "#,##0"; c.font = fnt(bold=True)
                c.fill = fl(bg); c.border = th(); c.alignment = ctr()

            elif key == "total_nok":
                formula = "+".join(f"{get_column_letter(next(i+1 for i,c2 in enumerate(cols) if c2['key']==k))}{r}" for k in nok_keys)
                if rw_ok_keys:
                    formula += "-" + "-".join(f"{get_column_letter(next(i+1 for i,c2 in enumerate(cols) if c2['key']==k))}{r}" for k in rw_ok_keys)
                c.value = f"={formula}"
                c.number_format = "#,##0"; c.font = fnt(bold=True)
                c.fill = fl("FFD0D0"); c.border = th(); c.alignment = ctr()

            elif key == "total_ok":
                c.value = f"={get_column_letter(total_col_ci())}{r}-{get_column_letter(tnok_col_ci())}{r}"
                c.number_format = "#,##0"; c.font = fnt(bold=True)
                c.fill = fl("D9F0D9"); c.border = th(); c.alignment = ctr()

            elif key == "nok_rate":
                c.value = f"=IFERROR({get_column_letter(tnok_col_ci())}{r}/{get_column_letter(total_col_ci())}{r},0)"
                c.number_format = "0.00%"
                c.font  = Font(name="Arial", bold=True, size=10, color="C00000" if has_nok else "000000")
                c.fill  = fl("FFF2CC" if has_nok else bg); c.border = th(); c.alignment = ctr()

            elif key == "remarks":
                c.value = ""; c.font = fnt(sz=9)
                c.fill = fl(bg); c.border = th(); c.alignment = lft()

            else:
                # Inspection or rework data column
                is_rework_col = any(key in [d["key_ok"], d["key_nok"]] for d in rework_list)
                is_nok_col = key.endswith("_nok")
                v = int(rd.get(key, 0))

                # For rework cols: only show for rework parts (multi-part) or always (single-part)
                if is_rework_col and multi_part and not is_rw_part:
                    c.value = ""
                    c.fill  = fl("F2F2F2")
                else:
                    c.value = v
                    c.number_format = "#,##0"
                    c.font  = Font(name="Arial",
                                   bold=(is_nok_col and v > 0),
                                   size=10,
                                   color="C00000" if (is_nok_col and v > 0) else "000000")
                    if is_rework_col:
                        c.fill = fl(C_RW_DATA)
                    elif is_nok_col:
                        c.fill = fl(C_NOK_DATA)
                    else:
                        c.fill = fl(C_OK_DATA)
                c.border = th(); c.alignment = ctr()

    # Total row
    total_row = DATA_START + len(by_day)
    ws.row_dimensions[total_row].height = 22
    label_span = min(2, static_count)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=label_span)
    ws.cell(row=total_row, column=1).value = "TOTAL"
    ws.cell(row=total_row, column=1).fill  = fl(C_TOTAL_ROW)
    ws.cell(row=total_row, column=1).font  = Font(name="Arial", bold=True, size=10)
    ws.cell(row=total_row, column=1).alignment = ctr()
    ws.cell(row=total_row, column=1).border = th()

    d0, d1 = DATA_START, total_row - 1
    for ci, col in enumerate(cols, 1):
        if ci <= label_span:
            continue
        c  = ws.cell(row=total_row, column=ci)
        cl = get_column_letter(ci)
        c.fill = fl(C_TOTAL_ROW); c.border = th()
        c.alignment = ctr(); c.font = Font(name="Arial", bold=True, size=10)
        key = col["key"]
        if key in ["total"] or (key.endswith("_ok") or key.endswith("_nok")):
            c.value = f"=SUM({cl}{d0}:{cl}{d1})"; c.number_format = "#,##0"
        elif key == "total_ok":
            c.value = f"={get_column_letter(total_col_ci())}{total_row}-{get_column_letter(tnok_col_ci())}{total_row}"
            c.number_format = "#,##0"
        elif key == "nok_rate":
            c.value = f"=IFERROR({get_column_letter(tnok_col_ci())}{total_row}/{get_column_letter(total_col_ci())}{total_row},0)"
            c.number_format = "0.00%"
            c.font = Font(name="Arial", bold=True, size=10, color="C00000")

    ws.freeze_panes = "A4"

    # ── SHEET 2: Summary by Part No. (multi-part only) ────────────────────────
    if multi_part and by_part is not None:
        ws2 = wb.create_sheet("Summary by Part No.")
        ws2.sheet_view.showGridLines = False

        s_cols = [{"key":"partno","label":"Part No.","width":18},
                  {"key":"total","label":"Total\nQty","width":12}]
        for ci2, d in enumerate(inspections):
            color = INSP_COLORS[ci2 % len(INSP_COLORS)]
            s_cols += [
                {"key":d["key_ok"],  "label":f"OK\n{d['label'][:14]}",  "width":12, "color":color},
                {"key":d["key_nok"], "label":f"NOK\n{d['label'][:14]}", "width":12, "color":color},
                {"key":f"rate_{d['key_nok']}", "label":"NOK\nRate", "width":10},
            ]
        for d in rework_list:
            s_cols += [
                {"key":d["key_ok"],  "label":"Rework\nOK",  "width":12, "color":RW_COLOR},
                {"key":d["key_nok"], "label":"Rework\nNOK", "width":12, "color":RW_COLOR},
            ]
        s_cols += [
            {"key":"total_ok",         "label":"Total\nOK",        "width":11},
            {"key":"total_nok",        "label":"Total\nNOK",       "width":11},
            {"key":"overall_nok_rate", "label":"Overall\nNOK Rate","width":13},
        ]
        ns = len(s_cols)
        for i, col in enumerate(s_cols, 1):
            ws2.column_dimensions[get_column_letter(i)].width = col["width"]

        ws2.row_dimensions[1].height = 32
        ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ns)
        c = ws2.cell(row=1, column=1)
        c.value = f"SUMMARY BY PART NUMBER  |  Order: {order_id}"
        c.font  = Font(name="Arial", bold=True, size=13, color="FFFFFF")
        c.fill  = fl(C_DARK); c.alignment = ctr()
        for ci2 in range(1, ns+1): ws2.cell(row=1, column=ci2).border = th()

        ws2.row_dimensions[2].height = 36
        for i, col in enumerate(s_cols, 1):
            c = ws2.cell(row=2, column=i)
            c.value = col["label"]; c.alignment = ctr(); c.border = th()
            key = col["key"]
            bg_c = col.get("color", C_MID)
            if "NOK" in col["label"] and "Rate" not in col["label"]:
                c.font = fnt(bold=True, sz=9); c.fill = fl(C_NOK_HDR)
            elif ("OK" in col["label"] and "NOK" not in col["label"]
                  and "Rate" not in col["label"] and "Overall" not in col["label"]):
                c.font = fnt(bold=True, sz=9); c.fill = fl(C_OK_HDR)
            else:
                c.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
                c.fill = fl(bg_c)

        insp_nok_keys = [d["key_nok"] for d in inspections]
        s_rw_ok_keys  = [d["key_ok"]  for d in rework_list]

        for idx, row in enumerate(by_part.itertuples(index=False)):
            r   = idx + 3
            ws2.row_dimensions[r].height = 18
            rd  = row._asdict()
            pno = rd["partno"]
            is_rw_part = pno in rework_parts
            t   = int(rd["total"])
            tnok = (sum(int(rd.get(k, 0)) for k in insp_nok_keys)
                    - sum(int(rd.get(k, 0)) for k in s_rw_ok_keys))
            bg  = C_HIGHLIGHT if is_rw_part else (C_ODD if idx%2==0 else C_WHITE)

            for ci2, col in enumerate(s_cols, 1):
                c = ws2.cell(row=r, column=ci2)
                key = col["key"]
                c.fill = fl(bg); c.border = th()
                c.alignment = lft() if ci2==1 else ctr()

                if key == "partno":
                    c.value = pno; c.font = fnt(bold=is_rw_part)
                elif key == "total":
                    c.value = t; c.number_format = "#,##0"; c.font = fnt(bold=True)
                elif key.startswith("rate_"):
                    nok_key = key[5:]
                    v = int(rd.get(nok_key, 0))
                    c.value = v/t if t > 0 else 0; c.number_format = "0.000%"
                    c.font  = Font(name="Arial", bold=v>0, size=10, color="C00000" if v>0 else "000000")
                elif key in [d["key_ok"] for d in rework_list] + [d["key_nok"] for d in rework_list]:
                    if is_rw_part:
                        v = int(rd.get(key, 0))
                        c.value = v; c.number_format = "#,##0"
                        is_bad = key.endswith("_nok") and v > 0
                        c.font  = Font(name="Arial", bold=is_bad, size=10,
                                       color="C00000" if is_bad else "000000")
                        c.fill = fl(C_RW_DATA)
                    else:
                        c.value = ""; c.fill = fl("F2F2F2")
                elif key == "total_ok":
                    c.value = t - tnok; c.number_format = "#,##0"
                    c.font  = fnt(bold=True); c.fill = fl("D9F0D9")
                elif key == "total_nok":
                    c.value = tnok; c.number_format = "#,##0"
                    c.font  = Font(name="Arial", bold=True, size=10,
                                   color="C00000" if tnok>0 else "000000")
                    c.fill = fl("FFD0D0")
                elif key == "overall_nok_rate":
                    c.value = tnok/t if t > 0 else 0; c.number_format = "0.000%"
                    c.font  = Font(name="Arial", bold=True, size=10,
                                   color="C00000" if tnok>0 else "000000")
                else:
                    v = int(rd.get(key, 0))
                    c.value = v; c.number_format = "#,##0"
                    is_nok  = key.endswith("_nok")
                    c.font  = Font(name="Arial", bold=(is_nok and v>0), size=10,
                                   color="C00000" if (is_nok and v>0) else "000000")
                    if key.endswith("_ok"):  c.fill = fl(C_OK_DATA)
                    elif is_nok:             c.fill = fl(C_NOK_DATA)

        # Summary total row
        sr  = len(by_part) + 3
        sd0, sd1 = 3, sr - 1
        ws2.row_dimensions[sr].height = 22
        for ci2 in range(1, ns+1):
            c = ws2.cell(row=sr, column=ci2)
            c.fill = fl(C_TOTAL_ROW); c.border = th()
            c.alignment = ctr(); c.font = Font(name="Arial", bold=True, size=10)
        ws2.cell(row=sr, column=1).value = "TOTAL"
        tot_ci = next(i+1 for i,c in enumerate(s_cols) if c["key"]=="total")
        tnok_ci= next(i+1 for i,c in enumerate(s_cols) if c["key"]=="total_nok")
        for ci2, col in enumerate(s_cols, 1):
            if ci2 == 1: continue
            c  = ws2.cell(row=sr, column=ci2)
            cl = get_column_letter(ci2)
            key = col["key"]
            if key in ["total","total_ok","total_nok"] or \
               (key.endswith("_ok") or key.endswith("_nok")):
                c.value = f"=SUM({cl}{sd0}:{cl}{sd1})"; c.number_format = "#,##0"
            elif key.startswith("rate_") or key == "overall_nok_rate":
                c.value = f"=IFERROR({get_column_letter(tnok_ci)}{sr}/{get_column_letter(tot_ci)}{sr},0)"
                c.number_format = "0.000%"
                c.font = Font(name="Arial", bold=True, size=10, color="C00000")
        ws2.freeze_panes = "B3"

    out = io.BytesIO()
    wb.save(out); out.seek(0)
    return out


# ── Streamlit UI ──────────────────────────────────────────────────────────────
uploaded = st.file_uploader(
    "Upload Awitec inspection file (.xlsx)",
    type=["xlsx"],
    help="Must contain a sheet named 'Order Awitec'"
)

if uploaded:
    with st.spinner("Reading file..."):
        info, err = parse_file(uploaded)

    if err:
        st.error(f"❌ {err}")
    else:
        meta = info["meta"]
        st.success("✅ File read successfully")

        c1, c2 = st.columns(2)
        with c1:
            st.markdown("**Order Details**")
            st.write(f"- Order: `{meta.get('order','–')}`")
            st.write(f"- Material: `{meta.get('material','–')}`")
            st.write(f"- Part numbers: **{len(info['part_numbers'])}**")
            st.write(f"- Total parts inspected: **{int(info['data']['total'].sum()):,}**")
        with c2:
            st.markdown("**Detected Inspections**")
            for d in info["inspections"]:
                nok = int(info["data"][d["key_nok"]].sum())
                st.write(f"- ✅ {d['name']} — NOK: **{nok:,}**")
            for d in info["rework"]:
                total_rw = int(info["data"][d["key_ok"]].sum() + info["data"][d["key_nok"]].sum())
                parts = ", ".join(info["rework_parts"]) if info["rework_parts"] else "all"
                st.write(f"- 🔧 {d['name']} ({parts}) — {total_rw:,} parts")
            if not info["inspections"]:
                st.warning("No inspections with data detected.")

        if info["part_numbers"]:
            with st.expander(f"Part numbers ({len(info['part_numbers'])})"):
                st.write(", ".join(info["part_numbers"]))

        st.divider()
        if st.button("Generate QC Report", type="primary", use_container_width=True):
            with st.spinner("Generating report..."):
                output = generate_report(info)
            order_match = re.search(r"(2600\d+)", uploaded.name)
            order_prefix = order_match.group(1) if order_match else meta.get("order", "QC").replace("/", "_").replace(" ", "_")
            fname = f"{order_prefix}_QC_Report.xlsx"
            st.download_button(
                label="⬇️  Download QC Report",
                data=output,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            st.success(f"Report ready: **{fname}**")
